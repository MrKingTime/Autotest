
import openpyxl
import re
import datetime
import pandas as pd
# class Ser:
#
#     def __init__(self):
#         self.name = 'abc'
#         self.date_msg = 'aawdawd'
#         self.keying = 'aklwdj'
#         self.itemAddress = '广东省广州市黄埔区科学城地铁旁创意大厦B3'
#         self.contractName = '创意大厦合同'
#         self.file_path = 'CN107415837A车载无人机和自动撑伞系统及方法.pdf'
#         self.excel_file_path = "data/casebase.xlsx"
#         self.module_name = "检测任务管理"
#
#
#     def execute_function(self,code):
#         return eval(code)
#
#
#     def process_excel(self):
#         """
#         Process Excel file to extract and process data under specified module_name.
#         """
#         data_dict = {}
#
#         # Load Excel file
#         wb = openpyxl.load_workbook(self.excel_file_path)
#         sheet = wb.active
#
#         # Find column indices for module name, case name, and case data
#         module_col = None
#         case_name_col = None
#         case_data_col = None
#         for col in range(1, sheet.max_column + 1):
#             cell_value = sheet.cell(row=1, column=col).value
#             if cell_value == "模块名称":
#                 module_col = col
#             elif cell_value == "用例名称":
#                 case_name_col = col
#             elif cell_value == "用例数据":
#                 case_data_col = col
#
#         if module_col is None or case_name_col is None or case_data_col is None:
#             print("Required columns not found in Excel sheet.")
#             return data_dict
#
#         # Find and process data under specified module name
#         for row in range(2, sheet.max_row + 1):
#             pass
#         module_value = sheet.cell(row=row, column=module_col).value
#         if module_value == self.module_name:
#             case_name = sheet.cell(row=row, column=case_name_col).value
#             case_data_str = sheet.cell(row=row, column=case_data_col).value
#             case_data = {}
#
#             # Process case data
#             matches = re.findall(r'\${(.*?)}', case_data_str)
#             for match in matches:
#                 # print(match)
#                 evaluated_value = self.execute_function(match)
#                 # print(evaluated_value)
#                 case_data_str = case_data_str.replace(f'${{{match}}}', str(evaluated_value))
#
#                 # Convert case data string to dictionary
#             try:
#                 case_data = eval(case_data_str)
#             except Exception as e:
#                 print(f"Error parsing case data for '{case_name}': {e}")
#
#             data_dict[case_name] = case_data
#         return data_dict

class Ser_excel():

    def __init__(self):
        self.name = 'abc'
        self.date_msg = 'aawdawd'
        self.itemAddress = '广东省广州市黄埔区科学城地铁旁创意大厦B3'
        self.contractName = '创意大厦合同'
        self.file_path = 'CN107415837A车载无人机和自动撑伞系统及方法.pdf'
        self.excel_file_path = "data/casebase.xlsx"

    def execute_function(self, code):
        return eval(code)

    def search(self):
        df = pd.read_excel(self.excel_file_path, sheet_name=0)
        case_data_str = df['用例数据'][0]
        matches = re.findall(r'\${(.*?)}', df['用例数据'][0])
        for match in matches:
            evaluated_value = self.execute_function(match)
            case_data_str = case_data_str.replace(f'${{{match}}}', str(evaluated_value))

        return eval(case_data_str)

